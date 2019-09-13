# First type command 'pip install openpyxl' in command prompt.

from openpyxl import load_workbook

def attendance_update():
    
#set a variable to the file path of our spreadsheet#
  #  excelfile = '##### LOCATION WHERE EXCEL FILE IS STORED#####'
    excelfile = 'E:\Ashutosh\.py Files\Project Email\simple_excelmail.xlsx'
#use the imported load_workbook() function to retrieve the spreadsheet#
    wb = load_workbook(excelfile)
#since we have the ability to store multiple spreadsheets in a single workbook, we’re setting up a #variable to access our first and only spreadsheet#
    ws = wb[wb.sheetnames[0]]
    
    attendance = []

    for row in ws.iter_rows(row_offset=1):  #If it causes error, try reemoving 'row_offset=1'
        for cell in row:
            print(cell.value)
            attendance.append(str(cell.value))
            print(type())
    return attendance

print(attendance_update())
