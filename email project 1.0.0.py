import smtplib

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook

'''
Requirements:
1. Open cmd and type: pip install openpyxl
2. Open google account of sender and allow unknown software to use gmail.
'''

gmail_user = "## senders email address ##"
gmail_pword = "## sender's password ##"

msg = MIMEMultipart('alternative')
sent_from = ['## senders email address ##']
msg['Subject'] = 'Attendance Status'
    
#excel-file=data

def excel_details():    
    
    excelfile = 'E:\simple_excelmail.xlsx'
    wb = load_workbook(excelfile)
    ws = wb[wb.sheetnames[0]]
    
    excel_list = []
    
    for row in ws.iter_rows(row_offset=1): # add 'row_offset=1' in ' () ' if there is an error.
        for cell in row:
#            print(cell.value)
            excel_list.append(str(cell.value))
            
    return excel_list

# details is Constructor. Builds template rows for html

class details:    
    
    def __init__(self, name, roll_no, date, attendance_per, status):
        self.name = name
        self.roll_no = roll_no.split(",")
        self.date = date.split(",")
        self.attendance_per = attendance_per.split(",")
        self.status = status.split(",")
        self.total = 0
        
        self.template_array = []
#        print(self.status[0], self.roll_no, self.to)
        
#        sum attendance_pers for total
        for i in range(0, len(self.roll_no)):
            self.total += int(self.attendance_per[i])
            
            template = """
                <tr class="left">
                    <td style="padding: 10px; text-align: left;">"""+ self.roll_no[i] +"""</td>
                    <td style="padding: 10px;">""" + self.date[i] + """</td>
                    <td style="text-align: right; padding: 10px;">"""+ self.attendance_per[i] +"""</td>
                    <td style="padding-left: 20px;">"""+ self.status[i] +"""</td>
                    </tr>
                """        
            self.template_array.append(template)
        self.total = str(self.total)

        
        
def send_email(to, msg, new_template):
        
    data.reverse()    
    
    for i in range(0, 6):
        data.pop()
           
    try:
        data.reverse()
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(gmail_user, gmail_pword)
        server.sendmail(sent_from, to, msg.as_string())
        
        print("Email Sent To: ", new_template.name)
        print("@: ", to)            
        print("Roll No Numbers: ", new_template.roll_no)
        print("TOTAL: ", new_template.total, "\n---------------------------")            
        server.quit()
        
        if(data[0] == 'None'):
            print('END OF LIST')
        else:            
            build_email(data)
    except Exception as e:
        print(e)
        print('Email Failed to Send to: ', new_template.name)
        print("@: ", to)
        print("details Numbers: ", new_template.roll_no)

data = excel_details()
#print(data)

def build_email(data):
    
#    print(data[5])
    new_template = details(data[0], data[1], data[2], data[3], data[4])
    
    to = data[5].split(",") #This must be a list i.e. ['someaddress@gmail.com']
    
    text = "Hello, {0}, roll no - {1} \n  As on the dates {2} , your  attendance  is {3}. You are encouraged to maintain minimum attendance of 75%.\n Thank You.\n TCY College.".format(data[0], data[1], data[2], data[3])
    
        #use IN-LINE Styling    
    html = """\
    <!DOCTYPE html>
    <html>
        <body>
            <p style="text-align: center"> Hello, """+ new_template.name +""" Hope this email finds you well.</p> 
              
            <p style="text-align: center">Here are your details.</p>
            <hr style="width: 500px;">
            <table style="margin-left: auto; margin-right: auto">
                <tr>
                    <th>Attendance TOTAL:</th>
                    <th style="padding-left: 100px">$"""+ new_template.total +"""</th>
                </tr>
            </table>
            <hr class="width">
            
            <table style="margin-left: auto; margin-right: auto">
                <tr class="left padded">
                    <th style="text-align: left;"> Roll No </th>
                    <th> Date </th>
                    <th style="text-align: right;"> attendance_per </th>
                    <th style="padding-left: 20px;"> status </th>
                </tr>
                """ + ''.join(new_template.template_array) + """
            </table>
            
            <hr style="width: 500px;">
                <table style="margin-left: auto; margin-right: auto; padding: 10px;">
                    <tr>
                        <th> Thank You! </th>
                    </tr>
                    <tr>
                        <th> TCY College </th>
                    </tr>
                </table>
            <hr style="width: 500px;">
            
        </body>
    </html>
    """
    

    part1 = MIMEText(text, 'plain')
    part2 = MIMEText(html, 'html')
    
    
    msg.attach(part1)
    msg.attach(part2)
      
    send_email(to, msg, new_template)  

                               
build_email(data)
